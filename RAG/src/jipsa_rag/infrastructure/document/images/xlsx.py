"""XLSX 삽입 이미지와 Microsoft Excel 차트 렌더 이미지를 추출한다.

삽입 이미지는 openpyxl의 drawing anchor를 기준으로 원본 시트와 셀 위치를
보존한다. 차트는 Excel 2024 COM의 ``Chart.Export``를 사용하여 각 차트를 직접
PNG로 출력한다. 전체 workbook을 PDF로 변환하지 않아 불필요한 페이지 렌더링과
임시 workbook 복제를 피한다.
"""

from __future__ import annotations

import asyncio
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

# openpyxl은 py.typed를 제공하지 않고, 외부 types-openpyxl 패키지는
# Workbook.active와 Chartsheet 타입을 실제 런타임보다 좁게 선언한다. 프로젝트
# 내부 타입 검사는 strict로 유지하되 이 외부 라이브러리 import 경계만 명시적으로
# 격리한다. 이후 동적 객체는 아래의 getattr, cast와 값 검증으로 좁힌다.
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from jipsa_rag.core.document_processing import DocumentProcessingSettings
from jipsa_rag.infrastructure.document.images.common import (
    build_image_id,
    can_append_image,
    guess_media_type,
    normalize_extension,
    validate_ooxml_image_package,
)
from jipsa_rag.infrastructure.document.images.models import (
    DocumentImageExtraction,
    DocumentImageKind,
    ExtractedDocumentImage,
)
from jipsa_rag.infrastructure.document.models import DocumentType
from jipsa_rag.infrastructure.document.rendering import (
    OfficeRenderClient,
    OfficeVisualRenderResult,
)

_XLSX_REQUIRED_MEMBERS = frozenset({"[Content_Types].xml", "xl/workbook.xml"})
_MAX_EXCEL_COLUMN = 16_384
_MAX_EXCEL_ROW = 1_048_576
_EMU_PER_PIXEL = 9_525
_DEFAULT_COLUMN_PIXELS = 64
_DEFAULT_ROW_PIXELS = 20


@dataclass(frozen=True, slots=True)
class _ChartContext:
    """한 XLSX 차트의 시트 위치와 OCR 주변 문맥."""

    sheet_index: int
    sheet_name: str
    chart_index: int
    anchor_cell: str
    cell_range: str
    context: str


class XlsxImageExtractor:
    """openpyxl 삽입 이미지와 Excel COM 차트 PNG를 결합한다."""

    def __init__(
        self,
        settings: DocumentProcessingSettings,
        renderer: OfficeRenderClient,
    ) -> None:
        self._settings = settings
        self._renderer = renderer

    @property
    def file_type(self) -> DocumentType:
        return DocumentType.XLSX

    async def extract(self, file_path: Path) -> DocumentImageExtraction:
        """OOXML을 검증한 뒤 삽입 이미지와 차트 렌더 이미지를 추출한다.

        Microsoft Excel은 문서를 실제로 여는 외부 응용 프로그램 경계다. ZIP 안전
        검증과 workbook 구조 분석을 먼저 완료하고, 차트가 확인된 문서에 한해서만
        COM 렌더러를 호출한다.
        """

        embedded_images, sheet_names, chart_contexts = await asyncio.to_thread(
            self._extract_embedded_sync,
            file_path,
        )

        if chart_contexts:
            rendered_result = await self._renderer.render_xlsx_charts(file_path)
        else:
            rendered_result = None

        rendered_images = self._build_rendered_images(
            rendered_result=rendered_result,
            chart_contexts=chart_contexts,
            existing_images=embedded_images,
            maximum_count=max(
                self._settings.image_max_count_per_document - len(embedded_images),
                0,
            ),
            image_index_offset=len(embedded_images),
        )
        images = tuple((*embedded_images, *rendered_images))

        return DocumentImageExtraction(
            images=images,
            document_metadata={
                "sheet_count": len(sheet_names),
                "extracted_image_count": len(images),
                "xlsx_picture_count": sum(
                    image.kind is DocumentImageKind.XLSX_PICTURE for image in images
                ),
                "xlsx_chart_count": len(chart_contexts),
                "xlsx_chart_render_count": sum(
                    image.kind is DocumentImageKind.XLSX_CHART_RENDER for image in images
                ),
                "office_renderer": "microsoft_office_com",
                "office_renderer_available": (
                    rendered_result.renderer_available if rendered_result is not None else False
                ),
            },
        )

    def _extract_embedded_sync(
        self,
        file_path: Path,
    ) -> tuple[
        tuple[ExtractedDocumentImage, ...],
        tuple[str, ...],
        tuple[_ChartContext, ...],
    ]:
        """안전 검증 후 삽입 이미지와 차트 anchor 문맥을 읽는다."""

        validate_ooxml_image_package(
            file_path,
            file_type=self.file_type,
            required_members=_XLSX_REQUIRED_MEMBERS,
            settings=self._settings,
        )
        workbook = load_workbook(
            filename=str(file_path),
            read_only=False,
            data_only=False,
        )
        images: list[ExtractedDocumentImage] = []
        chart_contexts: list[_ChartContext] = []
        sheet_names = tuple(workbook.sheetnames)

        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                sheet_context = _sheet_context(worksheet)

                # openpyxl은 차트·이미지 전체 컬렉션의 공개 열거 API를 제공하지
                # 않는다. 내부 컬렉션은 읽기 전용으로 복사하며 원본 workbook을
                # 변경하거나 저장하지 않는다.
                worksheet_charts = tuple(cast(list[Any], getattr(worksheet, "_charts", [])))
                for chart_index, chart in enumerate(worksheet_charts, start=1):
                    if len(chart_contexts) >= self._settings.image_max_count_per_document:
                        break
                    anchor_cell = _chart_anchor_cell(chart)
                    cell_range = _chart_anchor_range(chart)
                    chart_contexts.append(
                        _ChartContext(
                            sheet_index=sheet_index,
                            sheet_name=worksheet.title,
                            chart_index=chart_index,
                            anchor_cell=anchor_cell,
                            cell_range=cell_range,
                            context=(_cell_range_context(worksheet, cell_range) or sheet_context),
                        )
                    )

                worksheet_images = tuple(cast(list[Any], getattr(worksheet, "_images", [])))
                for image_index, image in enumerate(worksheet_images, start=1):
                    if len(images) >= self._settings.image_max_count_per_document:
                        break

                    content = _read_openpyxl_image_bytes(image)
                    if content is None:
                        # 손상된 단일 drawing part는 다른 이미지와 셀 텍스트의
                        # 색인을 막지 않는다.
                        continue

                    extension = normalize_extension(getattr(image, "format", "png"))
                    width_px = _positive_int(getattr(image, "width", None))
                    height_px = _positive_int(getattr(image, "height", None))
                    if not can_append_image(
                        images,
                        content,
                        width_px=width_px,
                        height_px=height_px,
                        settings=self._settings,
                    ):
                        continue

                    anchor_cell = _image_anchor_cell(image)
                    context = _cell_neighborhood_context(worksheet, anchor_cell)
                    images.append(
                        ExtractedDocumentImage(
                            image_id=build_image_id(
                                "xlsx",
                                sheet_index,
                                image_index,
                                anchor_cell,
                                content=content,
                            ),
                            kind=DocumentImageKind.XLSX_PICTURE,
                            content=content,
                            media_type=guess_media_type(extension),
                            extension=extension,
                            width_px=width_px,
                            height_px=height_px,
                            source_metadata={
                                "sheet_index": sheet_index,
                                "sheet_name": worksheet.title,
                                "image_index": image_index,
                                "cell_range": anchor_cell,
                                "shape_path": (
                                    f"sheet:{worksheet.title}/image:{image_index}/"
                                    f"anchor:{anchor_cell}"
                                ),
                            },
                            context_current=context or sheet_context,
                        )
                    )
        finally:
            workbook.close()

        return tuple(images), sheet_names, tuple(chart_contexts)

    def _build_rendered_images(
        self,
        *,
        rendered_result: OfficeVisualRenderResult | None,
        chart_contexts: tuple[_ChartContext, ...],
        existing_images: tuple[ExtractedDocumentImage, ...],
        maximum_count: int,
        image_index_offset: int,
    ) -> tuple[ExtractedDocumentImage, ...]:
        """Excel이 출력한 차트 PNG를 openpyxl 시트 문맥과 결합한다."""

        if rendered_result is None or maximum_count <= 0:
            return ()

        contexts = {
            (context.sheet_index, context.chart_index): context for context in chart_contexts
        }
        images: list[ExtractedDocumentImage] = []

        for visual in rendered_result.visuals:
            if len(images) >= maximum_count:
                break
            if visual.kind is not DocumentImageKind.XLSX_CHART_RENDER:
                continue
            if not can_append_image(
                (*existing_images, *images),
                visual.content,
                width_px=visual.width_px,
                height_px=visual.height_px,
                settings=self._settings,
            ):
                continue

            sheet_index = _metadata_int(visual.source_metadata.get("sheet_index"))
            chart_index = _metadata_int(visual.source_metadata.get("chart_index"))
            context = contexts.get((sheet_index, chart_index))
            source_metadata = dict(visual.source_metadata)
            if context is not None:
                # COM TopLeftCell은 Office 버전과 pywin32 동적 바인딩 상태에 따라
                # 실제 D15 차트를 A1로 반환할 수 있다. 시트·차트 순번은 COM 렌더 결과와
                # OOXML 구조가 동일하게 대응하므로, 위치 메타데이터는 원본 파일의
                # drawing anchor를 직접 읽은 openpyxl 값을 권위 있는 값으로 사용한다.
                office_anchor_cell = _metadata_text(source_metadata.get("anchor_cell"))
                if office_anchor_cell and office_anchor_cell != context.anchor_cell:
                    source_metadata["office_anchor_cell"] = office_anchor_cell

                source_metadata["sheet_name"] = context.sheet_name
                source_metadata["anchor_cell"] = context.anchor_cell
                source_metadata["cell_range"] = context.cell_range
                source_metadata["shape_path"] = (
                    f"sheet:{context.sheet_name}/chart:{chart_index}/anchor:{context.anchor_cell}"
                )

            images.append(
                ExtractedDocumentImage(
                    image_id=build_image_id(
                        "xlsx-office-render",
                        sheet_index,
                        chart_index,
                        content=visual.content,
                    ),
                    kind=DocumentImageKind.XLSX_CHART_RENDER,
                    content=visual.content,
                    media_type="image/png",
                    extension="png",
                    width_px=visual.width_px,
                    height_px=visual.height_px,
                    source_metadata={
                        **source_metadata,
                        "image_index": image_index_offset + len(images) + 1,
                    },
                    context_current=context.context if context is not None else "",
                )
            )

        return tuple(images)


def _read_openpyxl_image_bytes(image: Any) -> bytes | None:
    """openpyxl Image의 내부 데이터 함수를 제한된 경계에서 호출한다."""

    data_function = getattr(image, "_data", None)
    if not callable(data_function):
        return None
    try:
        return bytes(cast(Callable[[], bytes], data_function)())
    except Exception:
        return None


def _chart_anchor_cell(chart: Any) -> str:
    """openpyxl chart anchor의 정확한 왼쪽 위 셀을 A1 형식으로 반환한다.

    OCR 주변 문맥을 위한 ``cell_range``는 여백을 포함할 수 있지만, 출처 위치인
    ``anchor_cell``은 사용자가 Excel에서 지정한 실제 셀을 그대로 보존해야 한다.
    """

    anchor = getattr(chart, "anchor", None)
    if isinstance(anchor, str) and anchor:
        row, column = _coordinate_to_row_column(anchor)
        return f"{get_column_letter(column)}{row}"

    marker = getattr(anchor, "_from", None)
    if marker is None:
        return "A1"

    try:
        row = int(marker.row) + 1
        column = int(marker.col) + 1
    except (AttributeError, TypeError, ValueError):
        return "A1"

    return f"{get_column_letter(column)}{row}"


def _chart_anchor_range(chart: Any) -> str:
    """openpyxl chart anchor를 시트 문맥 조회용 셀 범위로 변환한다."""

    anchor = getattr(chart, "anchor", None)
    if isinstance(anchor, str) and anchor:
        start_row, start_column = _coordinate_to_row_column(anchor)
        end_row = min(start_row + 20, _MAX_EXCEL_ROW)
        end_column = min(start_column + 10, _MAX_EXCEL_COLUMN)
        return _format_cell_range(
            start_row,
            start_column,
            end_row,
            end_column,
        )

    marker = getattr(anchor, "_from", None)
    if marker is None:
        return "A1:L22"

    try:
        start_row = int(marker.row) + 1
        start_column = int(marker.col) + 1
    except (AttributeError, TypeError, ValueError):
        return "A1:L22"

    end_marker = getattr(anchor, "to", None)
    if end_marker is not None:
        try:
            end_row = int(end_marker.row) + 1
            end_column = int(end_marker.col) + 1
        except (AttributeError, TypeError, ValueError):
            end_row, end_column = _estimated_anchor_end(
                anchor,
                start_row=start_row,
                start_column=start_column,
            )
    else:
        end_row, end_column = _estimated_anchor_end(
            anchor,
            start_row=start_row,
            start_column=start_column,
        )

    # 한 셀의 여백을 포함하여 축 제목이나 데이터 레이블이 frame 경계에서
    # 잘리는 것을 줄인다.
    start_row = max(start_row - 1, 1)
    start_column = max(start_column - 1, 1)
    end_row = min(max(end_row + 1, start_row), _MAX_EXCEL_ROW)
    end_column = min(max(end_column + 1, start_column), _MAX_EXCEL_COLUMN)
    return _format_cell_range(
        start_row,
        start_column,
        end_row,
        end_column,
    )


def _estimated_anchor_end(
    anchor: Any,
    *,
    start_row: int,
    start_column: int,
) -> tuple[int, int]:
    """OneCellAnchor의 EMU 크기를 보수적인 셀 span으로 근사한다."""

    extent = getattr(anchor, "ext", None)
    if extent is None:
        width_pixels = 10 * _DEFAULT_COLUMN_PIXELS
        height_pixels = 20 * _DEFAULT_ROW_PIXELS
    else:
        try:
            width_pixels = max(int(extent.cx) // _EMU_PER_PIXEL, 1)
            height_pixels = max(int(extent.cy) // _EMU_PER_PIXEL, 1)
        except (AttributeError, TypeError, ValueError):
            width_pixels = 10 * _DEFAULT_COLUMN_PIXELS
            height_pixels = 20 * _DEFAULT_ROW_PIXELS

    column_span = max(
        (width_pixels + _DEFAULT_COLUMN_PIXELS - 1) // _DEFAULT_COLUMN_PIXELS,
        8,
    )
    row_span = max(
        (height_pixels + _DEFAULT_ROW_PIXELS - 1) // _DEFAULT_ROW_PIXELS,
        12,
    )
    return (
        min(start_row + row_span, _MAX_EXCEL_ROW),
        min(start_column + column_span, _MAX_EXCEL_COLUMN),
    )


def _format_cell_range(
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> str:
    return f"{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}"


def _coordinate_to_row_column(coordinate: str) -> tuple[int, int]:
    """A1 형식 좌표를 row, column 정수로 변환하고 실패 시 A1을 반환한다."""

    try:
        cell = coordinate.replace("$", "").split(":", maxsplit=1)[0]
        letters = "".join(character for character in cell if character.isalpha())
        digits = "".join(character for character in cell if character.isdigit())
        if not letters or not digits:
            raise ValueError

        column = 0
        for character in letters.upper():
            column = column * 26 + (ord(character) - ord("A") + 1)
        row = int(digits)
        if row < 1 or column < 1:
            raise ValueError
        return row, column
    except (TypeError, ValueError):
        return 1, 1


def _image_anchor_cell(image: Any) -> str:
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str) and anchor:
        return anchor
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return "A1"
    try:
        row = int(marker.row) + 1
        column = int(marker.col) + 1
    except (AttributeError, TypeError, ValueError):
        return "A1"
    return f"{get_column_letter(column)}{row}"


def _cell_neighborhood_context(worksheet: Any, anchor_cell: str) -> str:
    try:
        anchor = worksheet[anchor_cell]
        row = int(anchor.row)
        column = int(anchor.column)
    except Exception:
        return ""

    values: list[str] = []
    for row_index in range(max(1, row - 2), row + 3):
        row_values: list[str] = []
        for column_index in range(max(1, column - 2), column + 3):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is not None and str(value).strip():
                row_values.append(str(value).strip())
        if row_values:
            values.append(" | ".join(row_values))
    return "\n".join(values)


def _cell_range_context(worksheet: Any, cell_range: str) -> str:
    """차트 frame 주변 셀 값을 제한된 행·열 범위에서 직렬화한다."""

    start, _, end = cell_range.partition(":")
    start_row, start_column = _coordinate_to_row_column(start)
    end_row, end_column = _coordinate_to_row_column(end or start)
    end_row = min(end_row, start_row + 24)
    end_column = min(end_column, start_column + 14)

    values: list[str] = []
    for row_index in range(start_row, end_row + 1):
        row_values: list[str] = []
        for column_index in range(start_column, end_column + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is not None and str(value).strip():
                row_values.append(str(value).strip())
        if row_values:
            values.append(" | ".join(row_values))
    return "\n".join(values)


def _sheet_context(worksheet: Any) -> str:
    values: list[str] = []
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(worksheet.max_row, 20),
        min_col=1,
        max_col=min(worksheet.max_column, 12),
        values_only=True,
    ):
        normalized = [
            str(value).strip() for value in row if value is not None and str(value).strip()
        ]
        if normalized:
            values.append(" | ".join(normalized))
    return "\n".join(values)


def _positive_int(value: object) -> int | None:
    if isinstance(value, bool) or not isinstance(value, int | float):
        return None
    normalized = int(value)
    return normalized if normalized > 0 else None


def _metadata_text(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip()


def _metadata_int(value: object) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        return 0
    return value
