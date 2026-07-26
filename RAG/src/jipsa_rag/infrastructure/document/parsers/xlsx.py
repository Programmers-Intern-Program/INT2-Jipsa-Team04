"""XLSX 시트의 셀 값, 병합 범위, 표와 저장된 수식 결과를 추출한다.

XLSX는 표 계산 문서이므로 단순히 모든 셀을 이어 붙이면 행·열 위치와 수식
상태를 잃는다. 이 파서는 각 시트를 원본 행 단위로 순회하고, 실제 값이 존재하는
최소 열 범위를 탭 구분 문자열로 만든다. 각 행에는 다음 정보를 함께 기록한다.

- 시트 순번과 이름
- 원본 행 번호와 셀 범위
- Excel 표에 포함되는 행인지 여부와 표 이름
- 교차하는 병합 셀 범위
- 수식 셀 좌표와 수식 표현식
- Excel이 마지막 저장 시 기록한 수식 캐시 결과
- 캐시 결과가 존재하지 않는 수식 셀

``openpyxl``은 Excel 수식을 계산하는 엔진이 아니다. 따라서 이 파서는 동일 파일을
두 번 열어 ``data_only=False``에서는 수식 표현식을, ``data_only=True``에서는
저장된 캐시 결과를 읽는다. 캐시가 없으면 수식 표현식 자체를 검색 텍스트에 남기고,
메타데이터로 결과 누락 사실을 명시한다.
"""

import asyncio
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Final

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import MergedCell  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter, range_boundaries  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from jipsa_rag.infrastructure.document.exceptions import (
    DocumentParserError,
    DocumentReadError,
    DocumentTextExtractionError,
    DocumentTextNotFoundError,
    InvalidDocumentError,
)
from jipsa_rag.infrastructure.document.models import (
    DocumentType,
    ParsedDocument,
    ParsedDocumentUnit,
    SourceMetadataScalar,
    SourceMetadataValue,
)
from jipsa_rag.infrastructure.document.parsers._common import (
    normalize_text,
    validate_ooxml_package,
)

# 캐시된 수식 값을 우선 사용하는 XLSX 텍스트 파서임을 명시한다.
_XLSX_PARSER_TYPE: Final[str] = "XLSX_CACHED_VALUE"

# 1.1.0부터 시트 번호, 셀 시작/끝 좌표, 행·열 범위와 실제 값 셀 좌표를
# 명시적으로 보존한다. Parser_Version 변화는 Local RAG의 문서 식별자와
# 결정적 Chunk ID에 포함되므로 기존 1.0.0 색인을 자동으로 대체한다.
_XLSX_PARSER_VERSION: Final[str] = "1.1.0"

# ZIP 시그니처만으로는 다른 OOXML 형식과 구분할 수 없으므로 workbook 루트를 확인한다.
_XLSX_REQUIRED_MEMBERS: Final[frozenset[str]] = frozenset(
    {
        "[Content_Types].xml",
        "xl/workbook.xml",
    }
)


class XlsxDocumentParser:
    """``openpyxl``로 XLSX를 읽어 시트별 행 단위 검색 텍스트를 생성한다.

    하나의 비어 있지 않은 원본 행을 하나의 ``ParsedDocumentUnit``으로 만든다.
    이 경계는 표 행 비교와 원본 셀 위치 표시를 쉽게 하며, 한 청크가 서로 다른
    시트에 걸쳐 생성되는 것을 방지한다.
    """

    @property
    def file_type(self) -> DocumentType:
        """이 구현체가 처리하는 공통 문서 형식인 XLSX를 반환한다."""

        return DocumentType.XLSX

    @property
    def parser_type(self) -> str:
        """Local RAG DB에 저장할 XLSX 파서 종류를 반환한다."""

        return _XLSX_PARSER_TYPE

    @property
    def parser_version(self) -> str:
        """재파싱과 결정적 Chunk ID에 사용할 파서 호환 버전을 반환한다."""

        return _XLSX_PARSER_VERSION

    async def parse(self, file_path: Path) -> ParsedDocument:
        """동기식 Workbook 읽기와 셀 순회를 작업 스레드에서 실행한다."""

        return await asyncio.to_thread(self._parse_sync, file_path)

    def _parse_sync(self, file_path: Path) -> ParsedDocument:
        """XLSX 패키지를 열고 모든 시트를 원본 순서대로 파싱한다."""

        # 다운로드 계층 검증을 우회한 직접 호출도 안전하도록 XLSX 내부 루트를
        # 파서 경계에서 다시 확인한다.
        validate_ooxml_package(
            file_path,
            file_type=self.file_type,
            required_members=_XLSX_REQUIRED_MEMBERS,
        )

        try:
            # 첫 번째 Workbook은 수식 문자열을 보존한다. data_only=False일 때
            # formula_cell.value에는 "=SUM(A1:A3)" 같은 표현식이 들어 있다.
            formula_workbook = load_workbook(
                file_path,
                data_only=False,
                read_only=False,
                # 외부 통합 문서 링크는 검색 텍스트에 필요하지 않고, 외부 참조
                # 처리 복잡도를 줄이기 위해 로드하지 않는다.
                keep_links=False,
            )

            try:
                # 두 번째 Workbook은 Excel이 마지막 저장 시 기록한 캐시 결과를 읽는다.
                # openpyxl이 수식을 직접 계산하는 것이 아니라는 점에 주의해야 한다.
                value_workbook = load_workbook(
                    file_path,
                    data_only=True,
                    read_only=False,
                    keep_links=False,
                )
            except Exception:
                # 두 번째 로드 실패 시 먼저 열린 Workbook을 즉시 닫는다. Windows에서는
                # 열린 ZIP 핸들이 남아 있으면 임시 파일 삭제나 재시도가 실패할 수 있다.
                formula_workbook.close()
                raise

        except OSError as error:
            raise DocumentReadError(file_path) from error
        except Exception as error:
            # 잘못된 ZIP 관계, workbook XML 또는 지원하지 않는 손상 구조를
            # 공통 손상 문서 예외로 변환한다.
            raise InvalidDocumentError(self.file_type) from error

        try:
            units: list[ParsedDocumentUnit] = []
            table_names: list[str] = []
            table_ranges: list[str] = []
            merged_range_count = 0

            for sheet_index, formula_sheet in enumerate(
                formula_workbook.worksheets,
                start=1,
            ):
                # 두 Workbook은 같은 원본에서 열렸으므로 시트 이름으로 정확히 대응한다.
                value_sheet = value_workbook[formula_sheet.title]

                (
                    sheet_units,
                    sheet_table_names,
                    sheet_table_ranges,
                ) = self._parse_sheet(
                    formula_sheet,
                    value_sheet,
                    sheet_index=sheet_index,
                )

                units.extend(sheet_units)
                table_names.extend(sheet_table_names)
                table_ranges.extend(sheet_table_ranges)
                merged_range_count += len(formula_sheet.merged_cells.ranges)

            parsed_document = ParsedDocument(
                file_type=self.file_type,
                units=tuple(units),
                document_metadata={
                    "sheet_count": len(formula_workbook.worksheets),
                    "sheet_names": tuple(formula_workbook.sheetnames),
                    "table_names": tuple(table_names),
                    "table_ranges": tuple(table_ranges),
                    "merged_range_count": merged_range_count,
                    "source_unit_count": len(units),
                },
            )
        except DocumentParserError:
            raise
        except Exception as error:
            # 특정 행의 오류는 _parse_sheet()에서 위치별 예외로 바뀐다. 그 밖의
            # Workbook 관계나 시트 수준 오류는 유효하지 않은 문서로 처리한다.
            raise InvalidDocumentError(self.file_type) from error
        finally:
            # 성공과 실패 여부에 관계없이 두 Workbook의 ZIP 파일 핸들을 닫는다.
            formula_workbook.close()
            value_workbook.close()

        if parsed_document.text_unit_count == 0:
            raise DocumentTextNotFoundError(self.file_type)

        return parsed_document

    def _parse_sheet(
        self,
        formula_sheet: Worksheet,
        value_sheet: Worksheet,
        *,
        sheet_index: int,
    ) -> tuple[list[ParsedDocumentUnit], list[str], list[str]]:
        """단일 시트의 표·병합 범위를 준비하고 비어 있지 않은 행을 파싱한다."""

        # 행마다 모든 표와 병합 범위를 다시 탐색하면 큰 시트에서 불필요한 반복 비용이
        # 발생한다. 먼저 row_number별 역색인을 만들고 행 파싱에서 O(1)로 조회한다.
        table_by_row = self._map_tables_by_row(formula_sheet)
        merged_by_row = self._map_merged_ranges_by_row(formula_sheet)

        units: list[ParsedDocumentUnit] = []
        table_names = [table.displayName for table in formula_sheet.tables.values()]
        table_ranges = [table.ref for table in formula_sheet.tables.values()]

        for row_number in range(1, formula_sheet.max_row + 1):
            try:
                unit = self._parse_row(
                    formula_sheet,
                    value_sheet,
                    sheet_index=sheet_index,
                    row_number=row_number,
                    table_names=table_by_row.get(row_number, ()),
                    merged_ranges=merged_by_row.get(row_number, ()),
                )
            except DocumentParserError:
                raise
            except Exception as error:
                # 행 하나의 셀 값 변환이나 메타데이터 생성이 실패하면 정확한 시트와
                # 원본 행 번호를 포함한 추출 오류로 변환한다.
                raise DocumentTextExtractionError(
                    file_type=self.file_type,
                    source_metadata={
                        "sheet_name": formula_sheet.title,
                        "row_number": row_number,
                    },
                ) from error

            if unit is not None:
                # 완전히 빈 행은 unit을 만들지 않는다. 이후 행의 row_number는 실제
                # 원본 번호를 그대로 사용하므로 빈 행을 건너뛰어도 위치가 틀어지지 않는다.
                units.append(unit)

        return units, table_names, table_ranges

    def _parse_row(
        self,
        formula_sheet: Worksheet,
        value_sheet: Worksheet,
        *,
        sheet_index: int,
        row_number: int,
        table_names: tuple[str, ...],
        merged_ranges: tuple[str, ...],
    ) -> ParsedDocumentUnit | None:
        """하나의 원본 행을 탭 구분 문자열과 셀 위치 메타데이터로 변환한다."""

        meaningful_columns: list[int] = []
        meaningful_cells: list[str] = []
        display_values: dict[int, str] = {}

        # 수식 관련 정보는 같은 순번의 tuple끼리 대응하도록 동일 순서로 누적한다.
        formula_cells: list[str] = []
        formula_expressions: list[str] = []
        formula_results: list[SourceMetadataScalar] = []
        missing_formula_results: list[str] = []

        for column_number in range(1, formula_sheet.max_column + 1):
            formula_cell = formula_sheet.cell(
                row=row_number,
                column=column_number,
            )
            value_cell = value_sheet.cell(
                row=row_number,
                column=column_number,
            )

            # 병합 범위의 왼쪽 위 셀만 실제 값을 가진다. 나머지는 MergedCell 프록시이므로
            # 중복 텍스트와 읽기 전용 셀 문제를 피하기 위해 건너뛴다.
            if isinstance(formula_cell, MergedCell):
                continue

            formula_value = formula_cell.value
            cached_value = value_cell.value

            # openpyxl의 data_type="f"가 정상 신호지만 일부 생성 도구는 문자열만
            # "="로 시작하게 기록할 수 있어 두 조건을 모두 확인한다.
            is_formula = formula_cell.data_type == "f" or (
                isinstance(formula_value, str) and formula_value.startswith("=")
            )

            if formula_value is None and cached_value is None:
                continue

            meaningful_columns.append(column_number)
            meaningful_cells.append(formula_cell.coordinate)

            if is_formula:
                formula_cells.append(formula_cell.coordinate)
                formula_expressions.append(str(formula_value))
                formula_results.append(self._metadata_scalar(cached_value))

                if cached_value is None:
                    # LibreOffice, openpyxl 또는 일부 생성 도구가 저장한 파일에는 수식
                    # 캐시가 없을 수 있다. 누락 셀을 명시하여 "결과가 빈 값"과 구분한다.
                    missing_formula_results.append(formula_cell.coordinate)

            # 수식 캐시가 존재하면 사용자가 마지막으로 저장한 계산 결과를 검색 텍스트에
            # 사용한다. 캐시가 없으면 수식 표현식 자체를 남겨 정보가 완전히 사라지지 않게 한다.
            selected_value = (
                cached_value if is_formula and cached_value is not None else formula_value
            )
            display_values[column_number] = self._value_to_text(selected_value)

        if not meaningful_columns:
            return None

        first_column = min(meaningful_columns)
        last_column = max(meaningful_columns)

        # 첫 값과 마지막 값 사이의 빈 셀은 빈 문자열로 유지한다. 그래야 탭 개수가
        # 실제 열 간격을 반영하고, 중간 빈 셀 때문에 값이 왼쪽으로 이동하지 않는다.
        row_values = [
            display_values.get(column_number, "")
            for column_number in range(first_column, last_column + 1)
        ]

        cell_range = (
            f"{get_column_letter(first_column)}{row_number}:"
            f"{get_column_letter(last_column)}{row_number}"
        )

        start_cell = f"{get_column_letter(first_column)}{row_number}"
        end_cell = f"{get_column_letter(last_column)}{row_number}"

        metadata: dict[str, SourceMetadataValue] = {
            "unit_type": "table_row" if table_names else "row",
            "location_kind": "xlsx_cell_range",
            # sheet_index는 기존 계약과의 호환성을 유지한다. sheet_number를 같은
            # 1-based 값으로 추가하여 소비자가 인덱스의 0/1 시작 여부를 추측하지
            # 않아도 되게 한다.
            "sheet_index": sheet_index,
            "sheet_number": sheet_index,
            "sheet_name": formula_sheet.title,
            "row_number": row_number,
            "start_row": row_number,
            "end_row": row_number,
            "start_column": first_column,
            "end_column": last_column,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "cell_range": cell_range,
            "cell_coordinates": tuple(meaningful_cells),
            # cell_count는 실제 값 또는 수식이 있는 셀 수이고 column_span은
            # 중간 빈 셀까지 포함한 화면상의 열 범위 길이다.
            "cell_count": len(meaningful_columns),
            "column_span": last_column - first_column + 1,
            "merged_ranges": merged_ranges,
            "merged_cell_ranges": merged_ranges,
            "table_names": table_names,
            "formula_cells": tuple(formula_cells),
            "formula_expressions": tuple(formula_expressions),
            "formula_results": tuple(formula_results),
            "formula_result_missing_cells": tuple(missing_formula_results),
        }

        return ParsedDocumentUnit(
            text=normalize_text("\t".join(row_values).rstrip()),
            source_metadata=metadata,
        )

    @staticmethod
    def _map_tables_by_row(
        sheet: Worksheet,
    ) -> dict[int, tuple[str, ...]]:
        """시트의 Excel 표를 포함 행 번호별 표 이름 tuple로 역색인한다."""

        mutable_map: dict[int, list[str]] = {}

        for table in sheet.tables.values():
            # openpyxl 타입 정보가 불완전할 수 있으므로 실제 Table인지 확인한다.
            if not isinstance(table, Table):
                continue

            _, minimum_row, _, maximum_row = range_boundaries(table.ref)

            for row_number in range(minimum_row, maximum_row + 1):
                mutable_map.setdefault(row_number, []).append(table.displayName)

        return {row_number: tuple(names) for row_number, names in mutable_map.items()}

    @staticmethod
    def _map_merged_ranges_by_row(
        sheet: Worksheet,
    ) -> dict[int, tuple[str, ...]]:
        """병합 셀 범위를 교차하는 원본 행 번호별 문자열 tuple로 역색인한다."""

        mutable_map: dict[int, list[str]] = {}

        for merged_range in sheet.merged_cells.ranges:
            for row_number in range(
                merged_range.min_row,
                merged_range.max_row + 1,
            ):
                mutable_map.setdefault(row_number, []).append(str(merged_range))

        return {row_number: tuple(ranges) for row_number, ranges in mutable_map.items()}

    @staticmethod
    def _value_to_text(value: object) -> str:
        """셀 값을 결정적인 검색 문자열로 변환한다."""

        if value is None:
            return ""

        # 날짜·시간은 로컬 표시 형식보다 ISO 8601이 실행 환경에 독립적이고 검색 및
        # 비교가 쉽다. timezone 정보가 객체에 있으면 isoformat()이 함께 보존한다.
        if isinstance(value, datetime | date | time):
            return value.isoformat()

        # Python의 str(True)는 "True"지만 스프레드시트 표기와 일관되게 대문자를 쓴다.
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"

        return str(value)

    @staticmethod
    def _metadata_scalar(value: object) -> SourceMetadataScalar:
        """수식 캐시 결과를 공통 SourceMetadata 스칼라 타입으로 변환한다.

        SourceMetadata는 JSON 직렬화와 Qdrant payload 저장을 고려해 문자열, 정수,
        실수, 불리언과 null만 직접 허용한다. 날짜와 시간은 ISO 문자열로 바꾸고,
        Decimal은 float로 변환한다. 그 밖의 값은 안전한 문자열 표현으로 보존한다.
        """

        if value is None or isinstance(value, str | int | float | bool):
            return value

        if isinstance(value, Decimal):
            return float(value)

        if isinstance(value, datetime | date | time):
            return value.isoformat()

        return str(value)
